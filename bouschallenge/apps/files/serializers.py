from django.core.validators import FileExtensionValidator

from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import extend_schema_field
import openpyxl
from rest_framework import serializers, exceptions

from apps.entries.models import FileEntry
from apps.entries.serializers import CreateEntrySerializer, FileEntryListSerializer
from apps.files.models import UploadedFiles

allowed_columns = [
    'Cliente',
    '# Contrato',
    'Fecha de Compra',
    'Ciudad',
    'Empresa',
    'Valor adeudado',
]


class FileEntriesDetailsSerializer(serializers.ModelSerializer):
    entries = FileEntryListSerializer(many=True)

    class Meta:
        model = UploadedFiles
        fields = ('entries',)


class FileDetailsSerializer(serializers.ModelSerializer):
    entries = serializers.SerializerMethodField()
    total_debt = serializers.SerializerMethodField()
    debt_by_city = serializers.SerializerMethodField()
    debt_by_business = serializers.SerializerMethodField()

    class Meta:
        model = UploadedFiles
        fields = '__all__'
    
    @extend_schema_field(OpenApiTypes.INT)
    def get_entries(self, instance):
        return instance.entries.count()

    @extend_schema_field(OpenApiTypes.DECIMAL)
    def get_total_debt(self, instance):
        return instance.get_grand_total

    @extend_schema_field(OpenApiTypes.DECIMAL)
    def get_debt_by_city(self, instance):
        return instance.grand_debt_by_city

    @extend_schema_field(OpenApiTypes.DECIMAL)
    def get_debt_by_business(self, instance):
        return instance.grand_debt_by_business


class FileListSerializer(serializers.ModelSerializer):
    class Meta:
        model = UploadedFiles
        fields = '__all__'


class UploadSerializer(serializers.ModelSerializer):
    xcelfile = serializers.FileField(
        validators=[FileExtensionValidator(
            allowed_extensions=['xlsx', 'xls', 'xlt']
        )]
    )

    class Meta:
        model = UploadedFiles
        fields = ('xcelfile', 'filename',)
        extra_kwargs = {
            'filename': {'required': False},
        }
    
    def validate(self, attrs):
        filename = attrs['xcelfile'].name

        if UploadedFiles.objects.filter(filename=filename).exists():
            raise exceptions.ValidationError(
                {'xcelfile': 'Duplicated file'}
            )
        return attrs
    
    def create(self, validated_data):
        xcelfile = validated_data.get('xcelfile', '')
        file = UploadedFiles.objects.create(filename=xcelfile.name)
        wb = openpyxl.load_workbook(xcelfile)
        sheet = wb.active
        columns = {}
        # get columns
        for row in sheet.iter_rows(min_col=1, max_col=7):
            for cell in row:
                if cell.value is None:
                    continue
                if cell.value in allowed_columns:
                    columns[cell.value] = cell
        r_col = list()
        r_col.append(columns[list(columns)[0]].col_idx)
        r_col.append(columns[list(columns)[-1]].col_idx)
    
        for row in sheet.iter_rows(min_col=r_col[0], 
                                   max_col=r_col[-1], 
                                   min_row=columns[list(columns)[0]].row+1):
            row_data = dict()
            for cell in row:
                if cell._value is None:
                    break
                row_data[list(columns)[cell.col_idx-2]] = cell._value
            if len(row_data) > 0:
                newrow = dict()
                newrow['file'] = file.pk
                for item in row_data:
                    newrow[dict(FileEntry.COL_NAMES)[item]] = row_data[item]
                e = CreateEntrySerializer(data=newrow)
                if e.is_valid():
                    e.save()
        if len(file.entries.all()) == 0:
            file.delete()
            return False

        return file
